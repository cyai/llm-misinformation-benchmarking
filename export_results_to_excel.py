#!/usr/bin/env python3
"""
Export experiment results to Excel.

This script reads all iteration JSONL files from experiment results
and creates an Excel workbook with one sheet per strategy (zero_shot, one_shot, few_shot).

Each sheet contains all claims with their verdicts, confidence scores, rationales,
and other metadata across all iterations.

Usage:
    python export_results_to_excel.py --results-dir uku_expirements
    python export_results_to_excel.py --results-dir uku_expirements --output results_analysis.xlsx
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def load_jsonl(filepath):
    """Load JSONL file and return list of records."""
    records = []
    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError as e:
                    print(f"Warning: Skipping invalid JSON in {filepath}: {e}")
                    continue
    return records


def collect_all_results(results_dir):
    """
    Collect all results organized by strategy.

    Returns:
        dict: {strategy: list of all records from all iterations}
    """
    results_dir = Path(results_dir)

    # Find all strategy directories
    strategy_dirs = [d for d in results_dir.iterdir() if d.is_dir()]

    results_by_strategy = {}

    for strategy_dir in strategy_dirs:
        strategy_name = strategy_dir.name
        all_records = []

        # Load all iteration files
        iteration_files = sorted(strategy_dir.glob("iteration_*.jsonl"))

        for iteration_file in iteration_files:
            records = load_jsonl(iteration_file)
            all_records.extend(records)

        if all_records:
            results_by_strategy[strategy_name] = all_records
            print(f"Loaded {len(all_records)} records from {strategy_name}")

    return results_by_strategy


def create_dataframe(records):
    """
    Convert records to a pandas DataFrame with all relevant columns.

    Args:
        records: List of result dictionaries

    Returns:
        pd.DataFrame with organized columns
    """
    # Define column order
    columns = [
        "iteration",
        "sample_idx",
        "claim_id",
        "claim",
        "gold_label",
        "verdict",
        "confidence",
        "correct",
        "rationale",
        "cited_knowledge",
        "safety_notes",
        "timestamp",
        "reasoning_steps",  # For CoT
        "search_results",  # For search strategy
        "search_relevance",  # For search strategy
    ]

    # Create DataFrame
    df = pd.DataFrame(records)

    # Add 'correct' column - whether prediction matches gold label
    if "verdict" in df.columns and "gold_label" in df.columns:
        df["correct"] = df["verdict"] == df["gold_label"]

    # Reorder columns (only include columns that exist)
    existing_columns = [col for col in columns if col in df.columns]
    df = df[existing_columns]

    # Sort by iteration and sample_idx
    if "iteration" in df.columns and "sample_idx" in df.columns:
        df = df.sort_values(["iteration", "sample_idx"])

    return df


def style_worksheet(ws, strategy_name):
    """
    Apply styling to worksheet.

    Args:
        ws: openpyxl worksheet
        strategy_name: Name of the strategy for header
    """
    # Header style
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Apply header style to first row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Set column widths
    column_widths = {
        "A": 10,  # iteration
        "B": 12,  # sample_idx
        "C": 15,  # claim_id
        "D": 60,  # claim
        "E": 12,  # gold_label
        "F": 12,  # verdict
        "G": 12,  # confidence
        "H": 10,  # correct
        "I": 80,  # rationale
        "J": 50,  # cited_knowledge
        "K": 50,  # safety_notes
        "L": 20,  # timestamp
        "M": 80,  # reasoning_steps
        "N": 80,  # search_results
        "O": 15,  # search_relevance
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row and first two columns
    ws.freeze_panes = "C2"

    # Apply borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Data rows
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Color-code correct/incorrect predictions
    for row_idx in range(2, ws.max_row + 1):
        correct_cell = ws[f"H{row_idx}"]
        if correct_cell.value == True:
            correct_cell.fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            correct_cell.font = Font(color="006100")
        elif correct_cell.value == False:
            correct_cell.fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            correct_cell.font = Font(color="9C0006")


def create_summary_sheet(wb, results_by_strategy):
    """
    Create a summary sheet with aggregate statistics.

    Args:
        wb: openpyxl Workbook
        results_by_strategy: Dict of strategy -> records
    """
    ws = wb.create_sheet("Summary", 0)

    # Header
    ws["A1"] = "Experiment Results Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    # Column headers
    headers = [
        "Strategy",
        "Total Samples",
        "Correct",
        "Incorrect",
        "Accuracy (%)",
        "Avg Confidence",
        "Iterations",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.font = Font(bold=True, color="FFFFFF")

    # Calculate statistics for each strategy
    row = 4
    for strategy, records in sorted(results_by_strategy.items()):
        df = pd.DataFrame(records)

        total = len(df)
        if "verdict" in df.columns and "gold_label" in df.columns:
            correct = (df["verdict"] == df["gold_label"]).sum()
            incorrect = total - correct
            accuracy = (correct / total * 100) if total > 0 else 0
        else:
            correct = incorrect = accuracy = 0

        avg_confidence = df["confidence"].mean() if "confidence" in df.columns else 0
        iterations = df["iteration"].nunique() if "iteration" in df.columns else 0

        ws.cell(row=row, column=1, value=strategy)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=correct)
        ws.cell(row=row, column=4, value=incorrect)
        ws.cell(row=row, column=5, value=f"{accuracy:.2f}")
        ws.cell(row=row, column=6, value=f"{avg_confidence:.3f}")
        ws.cell(row=row, column=7, value=iterations)

        row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 12

    # Add borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_cells in ws.iter_rows(min_row=3, max_row=row - 1, min_col=1, max_col=7):
        for cell in row_cells:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def export_to_excel(results_dir, output_file):
    """
    Export all experiment results to Excel.

    Args:
        results_dir: Path to results directory
        output_file: Output Excel file path
    """
    print(f"Loading results from {results_dir}...")
    results_by_strategy = collect_all_results(results_dir)

    if not results_by_strategy:
        print("No results found!")
        sys.exit(1)

    print(f"\nCreating Excel workbook: {output_file}")
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create summary sheet
    create_summary_sheet(wb, results_by_strategy)

    # Create a sheet for each strategy
    for strategy, records in sorted(results_by_strategy.items()):
        print(f"Creating sheet for {strategy}...")

        # Convert to DataFrame
        df = create_dataframe(records)

        # Create worksheet
        ws = wb.create_sheet(title=strategy)

        # Write DataFrame to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

        # Apply styling
        style_worksheet(ws, strategy)

    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Excel file created: {output_file}")
    print(f"  - Summary sheet with aggregate statistics")
    print(f"  - {len(results_by_strategy)} strategy sheets with detailed results")


def main():
    parser = argparse.ArgumentParser(
        description="Export experiment results to Excel with one sheet per strategy"
    )
    parser.add_argument(
        "--results-dir",
        type=str,
        default="uku_expirements",
        help="Directory containing experiment results (default: uku_expirements)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output Excel file (default: <results_dir>_results.xlsx)",
    )

    args = parser.parse_args()

    # Set default output filename
    if args.output is None:
        results_dir_name = Path(args.results_dir).name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = f"{results_dir_name}_results_{timestamp}.xlsx"

    # Export
    export_to_excel(args.results_dir, args.output)


if __name__ == "__main__":
    main()
